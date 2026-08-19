import openpyxl

wb = openpyxl.load_workbook('大學部111-115成績_含折線圖.xlsx', data_only=False)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"=== Sheet: {sheetname} ===")
    print(f"Dimensions: {ws.dimensions}, Max row: {ws.max_row}, Max col: {ws.max_column}")
    print("Row 1:", [c.value for c in ws[1]])
    print("Row 3:", [c.value for c in ws[3]])
    print("Charts count:", len(ws._charts))
    if ws._charts:
        chart = ws._charts[0]
        print(f"Chart title: {chart.title}, type: {type(chart)}, style: {chart.style}")
        print(f"Chart width: {chart.width}, height: {chart.height}")
