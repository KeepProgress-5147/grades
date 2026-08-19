import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import json
import re
import os

def calculate_competition_ranks(scores):
    """
    scores: list of (index, score_val) where score_val can be float or None.
    returns dict: index -> rank_int or None
    Standard competition ranking (1224)
    """
    valid = [(idx, val) for idx, val in scores if val is not None]
    if not valid:
        return {idx: None for idx, _ in scores}
    
    # Sort descending by score
    valid.sort(key=lambda x: x[1], reverse=True)
    
    ranks = {}
    current_rank = 1
    for i, (idx, val) in enumerate(valid):
        if i > 0 and val == valid[i-1][1]:
            ranks[idx] = ranks[valid[i-1][0]]
        else:
            ranks[idx] = i + 1
            
    for idx, val in scores:
        if val is None:
            ranks[idx] = None
            
    return ranks

def update_all():
    print("Reading 大學部111-115成績.xlsx...")
    wb_in = openpyxl.load_workbook('大學部111-115成績.xlsx', data_only=True)
    ws_in = wb_in.active

    # Parse categorized items
    groups_data = {}
    
    # Group order to maintain consistency
    group_order = ['醫學', '護理', '聽語', '視光', '醫檢']
    
    for r in ws_in.iter_rows(min_row=2, values_only=True):
        school = r[0]
        dept = r[1]
        group = r[2]
        if not school or not dept or not group or not str(group).strip():
            continue
        g = str(group).strip()
        if g not in groups_data:
            groups_data[g] = []
            if g not in group_order:
                group_order.append(g)
                
        scores = []
        for v in [r[3], r[4], r[5], r[6], r[7]]:
            if v is not None:
                try:
                    scores.append(round(float(v), 2))
                except:
                    scores.append(None)
            else:
                scores.append(None)
                
        groups_data[g].append({
            'school': str(school).strip(),
            'dept': str(dept).strip(),
            'group': g,
            'scores': scores
        })

    # Calculate ranks for each group
    json_raw_data = {}
    summary_rows = []

    for g in group_order:
        if g not in groups_data:
            continue
        items = groups_data[g]
        json_raw_data[g] = []
        
        # Calculate ranks across 5 years
        # for each year (0 to 4):
        year_ranks = []
        for y_idx in range(5):
            year_scores = [(i, item['scores'][y_idx]) for i, item in enumerate(items)]
            r_map = calculate_competition_ranks(year_scores)
            year_ranks.append(r_map)
            
        for i, item in enumerate(items):
            ranks = [year_ranks[y_idx][i] for y_idx in range(5)]
            item['ranks'] = ranks
            
            # minScore, maxScore, diff
            valid_scores = [s for s in item['scores'] if s is not None]
            min_score = min(valid_scores) if valid_scores else None
            max_score = max(valid_scores) if valid_scores else None
            
            if item['scores'][4] is not None and item['scores'][0] is not None:
                diff = round(item['scores'][4] - item['scores'][0], 2)
            else:
                diff = None
            item['diff'] = diff
            
            is_public = item['school'].startswith('國立') or item['school'].startswith('市立') or '公立' in item['school']
            
            json_raw_data[g].append({
                'school': item['school'],
                'dept': item['dept'],
                'fullName': f"{item['school']} {item['dept']}",
                'isPublic': is_public,
                'scores': item['scores'],
                'ranks': item['ranks'],
                'minScore': min_score,
                'maxScore': max_score,
                'diff': diff
            })
            
            # For summary sheet in Excel
            summary_rows.append({
                'school': item['school'],
                'dept': item['dept'],
                'group': g,
                'scores': item['scores'],
                'ranks': item['ranks'],
                'diff': diff
            })

    print(f"Total processed items across {len(groups_data)} groups: {len(summary_rows)}")

    # 1. Generate 大學部111-115成績_含折線圖.xlsx
    wb_out = openpyxl.Workbook()
    
    # Styles
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font = Font(name="微軟正黑體", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="微軟正黑體", size=10)
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Sheet 1: 總表彙整
    ws_summary = wb_out.active
    ws_summary.title = "總表彙整"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title
    ws_summary.merge_cells("A1:N1")
    t_cell = ws_summary["A1"]
    t_cell.value = "大學部 111-115 學年度各群組錄取平均分數與排名彙整表"
    t_cell.fill = title_fill
    t_cell.font = title_font
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 36
    ws_summary.row_dimensions[2].height = 10

    headers_summary = [
        "校名", "系組名", "歸類群組",
        "111年分數", "111排名",
        "112年分數", "112排名",
        "113年分數", "113排名",
        "114年分數", "114排名",
        "115年分數", "115排名",
        "五年分數增減"
    ]
    ws_summary.append([]) # row 2 empty
    ws_summary.append(headers_summary) # row 3 headers
    ws_summary.row_dimensions[3].height = 26

    for col_idx in range(1, 15):
        c = ws_summary.cell(row=3, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    row_idx = 4
    for r in summary_rows:
        row_vals = [
            r['school'],
            r['dept'],
            r['group'],
            r['scores'][0], f"#{r['ranks'][0]}" if r['ranks'][0] is not None else "-",
            r['scores'][1], f"#{r['ranks'][1]}" if r['ranks'][1] is not None else "-",
            r['scores'][2], f"#{r['ranks'][2]}" if r['ranks'][2] is not None else "-",
            r['scores'][3], f"#{r['ranks'][3]}" if r['ranks'][3] is not None else "-",
            r['scores'][4], f"#{r['ranks'][4]}" if r['ranks'][4] is not None else "-",
            r['diff'] if r['diff'] is not None else "-"
        ]
        ws_summary.append(row_vals)
        ws_summary.row_dimensions[row_idx].height = 20
        
        # Alignment & Formats
        ws_summary.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws_summary.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws_summary.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
        
        # 111-115 Scores & Ranks
        for y_i in range(5):
            s_col = 4 + y_i * 2
            r_col = 5 + y_i * 2
            sc_cell = ws_summary.cell(row=row_idx, column=s_col)
            sc_cell.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(sc_cell.value, (int, float)):
                sc_cell.number_format = "#,##0.00"
            ws_summary.cell(row=row_idx, column=r_col).alignment = Alignment(horizontal="center", vertical="center")
            
        diff_cell = ws_summary.cell(row=row_idx, column=14)
        diff_cell.alignment = Alignment(horizontal="right", vertical="center")
        if isinstance(diff_cell.value, (int, float)):
            diff_cell.number_format = "+#,##0.00;-#,##0.00;0.00"

        for col_idx in range(1, 15):
            c = ws_summary.cell(row=row_idx, column=col_idx)
            c.font = data_font
            c.border = thin_border
            
        row_idx += 1

    # Adjust width for Summary
    for col in ws_summary.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            char_len = sum(2 if ord(ch) > 127 else 1 for ch in val_str)
            if char_len > max_len:
                max_len = char_len
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # Group Sheets
    for g in group_order:
        if g not in groups_data:
            continue
        items = groups_data[g]
        ws_g = wb_out.create_sheet(title=f"{g}組")
        ws_g.views.sheetView[0].showGridLines = True
        
        # Title
        ws_g.merge_cells("A1:F1")
        gt_cell = ws_g["A1"]
        gt_cell.value = f"【{g}組】111-115學年度歷年平均分數、組內排名統計與折線圖"
        gt_cell.fill = title_fill
        gt_cell.font = title_font
        gt_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_g.row_dimensions[1].height = 36
        ws_g.row_dimensions[2].height = 10

        g_headers = ["校系名稱 (圖例)", "111年", "112年", "113年", "114年", "115年"]
        ws_g.append([]) # row 2
        ws_g.append(g_headers) # row 3
        ws_g.row_dimensions[3].height = 24
        
        for col_idx in range(1, 7):
            c = ws_g.cell(row=3, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

        g_row_idx = 4
        for item in items:
            row_vals = [
                f"{item['school']} {item['dept']}",
                item['scores'][0],
                item['scores'][1],
                item['scores'][2],
                item['scores'][3],
                item['scores'][4]
            ]
            ws_g.append(row_vals)
            ws_g.row_dimensions[g_row_idx].height = 20
            
            ws_g.cell(row=g_row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
            for c_i in range(2, 7):
                sc = ws_g.cell(row=g_row_idx, column=c_i)
                sc.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(sc.value, (int, float)):
                    sc.number_format = "#,##0.00"
            for col_idx in range(1, 7):
                c = ws_g.cell(row=g_row_idx, column=col_idx)
                c.font = data_font
                c.border = thin_border
            g_row_idx += 1

        # Adjust column widths for Group sheet
        for col in ws_g.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                char_len = sum(2 if ord(ch) > 127 else 1 for ch in val_str)
                if char_len > max_len:
                    max_len = char_len
            ws_g.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Add Line Chart
        chart = LineChart()
        chart.title = f"{g}組 111-115學年度歷年平均分數走勢"
        chart.style = 13
        chart.y_axis.title = "平均分數"
        chart.x_axis.title = "學年度"
        chart.width = 15
        chart.height = 8.5
        
        # Data reference: columns B to F (2 to 6), rows 3 to max_row
        data_ref = Reference(ws_g, min_col=1, min_row=3, max_col=6, max_row=g_row_idx-1)
        chart.add_data(data_ref, titles_from_data=True, from_rows=True)
        
        # Categories reference: row 3, cols 2 to 6
        cats = Reference(ws_g, min_col=2, max_col=6, min_row=3, max_row=3)
        chart.set_categories(cats)
        
        ws_g.add_chart(chart, "H3")

    wb_out.save("大學部111-115成績_含折線圖.xlsx")
    print("Saved: 大學部111-115成績_含折線圖.xlsx")

    # 2. Update index.html and 大學部111-115成績_互動圖表.html
    raw_data_json_str = json.dumps(json_raw_data, ensure_ascii=False, indent=2)
    
    for html_file in ['index.html', '大學部111-115成績_互動圖表.html']:
        if not os.path.exists(html_file):
            continue
        with open(html_file, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
        # Replace const rawData = {...};
        pattern = re.compile(r'const rawData\s*=\s*\{.*?\};', re.DOTALL)
        new_raw_data_str = f"const rawData = {raw_data_json_str};"
        
        if pattern.search(html_content):
            updated_html = pattern.sub(new_raw_data_str, html_content, count=1)
            with open(html_file, 'w', encoding='utf-8') as f:
                f.write(updated_html)
            print(f"Updated {html_file} with latest rawData JSON!")
        else:
            print(f"Warning: pattern not matched in {html_file}")

if __name__ == "__main__":
    update_all()
